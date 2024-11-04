import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
import zipfile
import datetime

# Load environment variables
load_dotenv()

def send_weather_email(weather_data, averages, recipient_emails):
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = os.getenv("EMAIL_ADDRESS")
    sender_password = os.getenv("EMAIL_PASSWORD")

    # Compose the email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(recipient_emails)
    msg['Subject'] = "Weather Report"

    # Format the weather data into an HTML table
    body = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
         <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            h1 { color: #333; }
            h2 { color: #555; margin-top: 20px; }
            table { width: 80%; border-collapse: collapse; margin-bottom: 20px; }
            th, td { border: 1px solid #ddd; padding: 10px; text-align: left; }
            th { background-color: #f2f2f2; color: #333; }
            tr:nth-child(even) { background-color: #f9f9f9; }
            tr:hover { background-color: #ddd; }
            /* Specific styles for averages */
            .averages-table {
                width: 80%;
                margin-top: 20px;
                border: 2px solid #007BFF; /* blue border for emphasis */
            }
            .averages-table th {
                background-color: #007BFF; /* blue background for headers */
                color: white; /* white text for headers */
                font-weight: bold;
            }
            .averages-table td {
                background-color: #E7F1FF; /* light blue background for data */
            }
        </style>
    </head>
    <body>
    <h1>Weather Report</h1>
    """
    body = printAverages(averages, body)

    for source, data in weather_data.items():
        body += f"<h2>Source: {source}</h2>"
        body += "<table border='1'><tr>"
        
        # Add table headers from the keys of the first item in the list
        headers = data[0].keys()
        for key in headers:
            body += f"<th>{key.capitalize()}</th>"
        body += "</tr>"

        # Populate the table rows with weather data
        for city_data in data:
            body += "<tr>"
            for key in headers:
                body += f"<td>{city_data[key]}</td>"
            body += "</tr>"
        body += "</table>"

    # Close the HTML tags
    body += "</body>"

    msg.attach(MIMEText(body, 'html'))  # Attach the HTML body

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Secure the connection
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipient_emails, msg.as_string())
        print("Email sent successfully.")
        log_email_send("Success", body)  # Log success if email is sent
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        log_email_send("Failed", body)  # Log failure if email is not sent
        return False

def log_email_send(status="Failed", message="message that was sent, by email", log_file="Log.xlsx"):
    try:
        wb = load_workbook(log_file)  # Attempt to load existing workbook
        sheet = wb.active
    except (FileNotFoundError, zipfile.BadZipFile):  # Handle both file not found and bad zip file
        wb = Workbook()  # Create a new workbook
        sheet = wb.active
        sheet.append(["Timestamp", "Status", "Message"])  # Add headers

    # Append log entry
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([timestamp, status, message])  # Log the email send status

    wb.save(log_file)  # Save workbook
    print("Logged email send to Log.xlsx")

def printAverages(averages, body):
    emailBody = body
    emailBody += """
    <h2>Averages:</h2>
    <table class="averages-table">
    <tr>
    """
    for key in averages[0].keys():
        emailBody += f"<th>{key.capitalize()}</th>"
    emailBody += "</tr>"
    for average in averages:
        emailBody += "<tr>"
        for key in average.keys():
            emailBody += f"<td>{average[key]}</td>"
        emailBody += "</tr>"
    emailBody += "</table>"
    return emailBody