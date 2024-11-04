import os
import pandas
from openpyxl import load_workbook
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv

from weather_bot import log_email_send

load_dotenv()

#Default admins
defaultAdmins = os.getenv("DEFAULT_ADMINS")

errorLogFilePath = "error_log.xlsx";

#defines error_info
error_info = {
    "errLVL": "Error",
    "errLocation": "tasks.py, line 25",
    "errMsg": "File not found",
}


def reportErrorData(errorLVL,error,tb):
    """
        Used for logging errors in the application

        Parameters:
            errorLVL:
                Should be a string that indicates how critical an error is
            error:
                should be the error object that containes the error message, can also be a string containing the error message
            tb:
                TB is used for getting information on where the error occured. This is what should be given to this part excluding the quotation marks(""), "traceback.extract_tb(e.__traceback__)"
    """
    error_info["errLVL"] = str(errorLVL);
    error_info["errMsg"] = str(error)        
    
    if(isinstance(tb,str)):
        error_info["errLocation"] = tb;
    else:
        # Use traceback to capture the file name and line number
        file_path, line_number = tb[-1].filename, tb[-1].lineno

        # Extract just the file name
        file_name = os.path.basename(file_path)
        error_info["errLocation"] = f"{file_name}, line {line_number}"     
            
    __logErrorToExcel(error_info)

def __logErrorToExcel(error_details={"errMsg":"No message defined","errLVL":"unknown", "errLocation":"unknown file and row"}):
    """
    Write error message in error_log.xlsx
    Parameters:
        error_details should contain all the data that is needed for logging the info into excel.
        This would include the following:
            errMsg:
                message containing a description of the error
            errLVL:
                how critical is the error
            errLocation:
                where, or what file did the error occur in

    # Example usage
    error_info = {
        "Error Level": "Error",
        "Location": "data_processing.py, line 54",
        "Error Message": "File not found",
    }
    logErrorToExcel(error_info)
    """
    try:
        workbook = load_workbook(errorLogFilePath)
        sheet = workbook.active
    except FileNotFoundError:
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        # Write headers if file is new
        sheet.append(["Timestamp", "Error Level", "Location","Error Message"])
    # Append error details as a new row
    currentTime = datetime.now().strftime("%Y-%m-%d %H:%M:%S");
    sheet.append([
        currentTime,
        error_details.get("errLVL", ""),
        error_details.get("errLocation",""),
        error_details.get("errMsg", ""),
    ])

    # Save the workbook
    workbook.save(errorLogFilePath)

    errorContent = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            table {{
                font-family: Arial, sans-serif;
                border-collapse: collapse;
                width: 100%;
            }}
            th, td {{
                border: 1px solid #dddddd;
                text-align: left;
                padding: 8px;
            }}
            th {{
                background-color: #f2f2f2;
            }}
        </style>
    </head>
    <body>

    <h2>Error Log</h2>

    <table>
        <tr>
            <th>Timestamp</th>
            <td>{currentTime}</td>
        </tr>
        <tr>
            <th>Error Level</th>
            <td>{error_details.get("errLVL", "")}</td>
        </tr>
        <tr>
            <th>Location</th>
            <td>{error_details.get("errLocation", "")}</td>
        </tr>
        <tr>
            <th>Error Message</th>
            <td>{error_details.get("errMsg", "")}</td>
        </tr>
    </table>

    </body>
    </html>
    """
    __notifyAdministrator(errorContent)

def __notifyAdministrator(body):
    print("Notifying...")

    admin_emails = __getAdministrators();

    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = os.getenv("EMAIL_ADDRESS")
    sender_password = os.getenv("EMAIL_PASSWORD")


    if admin_emails is []:
        admin_emails = defaultAdmins
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ", ".join(admin_emails)
        msg['Subject'] = "No administrators found in Users.xlsx"
        body = "The Users.xlsx file does not contain any administrators. Default administrators have been notified."
        msg.attach(MIMEText(body, 'plain'))
        try:
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()  # Secure the connection
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, admin_emails, msg.as_string())
            print("Email sent successfully.")
            log_email_send("Success", body)  # Log success if email is sent
        except Exception as e:
            print(f"Failed to send email: {e}")
            log_email_send("Failed", body)  # Log failure if email is not sent

    # Compose the email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(admin_emails)
    msg['Subject'] = "There has been an error, while making Weather Report"
    msg.attach(MIMEText(body, 'html'))
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Secure the connection
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, admin_emails, msg.as_string())
        print("Email sent successfully.")
        log_email_send("Success", body)  # Log success if email is sent
    except Exception as e:
        print(f"Failed to send email: {e}")
        log_email_send("Failed", body)  # Log failure if email is not sent

def __getAdministrators():
    """
    Gets a list of all administrator emails
    """
    print("Getting administrators...")

    try:
            # Read the Excel file into a DataFrame
            df = pandas.read_excel("Users.xlsx")
            
            # Check the columns (for reference)
            print("Columns in the file:", df.columns.tolist())

            # Filter rows where the Role is "administrator" (case insensitive)
            admins = df[df["Role"].str.lower() == "administrator"]

            # Extract the email addresses from the filtered DataFrame
            admin_emails = admins["User List"].tolist()

            # Check if any administrators were found
            if not admin_emails:
                print("No administrators found.")
            
            return admin_emails

    except FileNotFoundError:
        print("Error: 'Users.xlsx' file not found.")
        return []
    except KeyError as e:
        print(f"Error: Missing expected column in the spreadsheet: {e}")
        return []
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return []